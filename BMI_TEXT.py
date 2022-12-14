#Nowy projekt. IKSDE

from openpyxl import load_workbook
wb = load_workbook("bmi_tabela.xlsx")
sheet = wb.active

#for row in sheet.iter_rows(min_row=3, min_col=2, max_row=22, max_col=5):
#   for cell in row:
#        print(cell.value)
        
ah = sheet["D3"].value
aw = sheet["E3"].value
aBMI = (aw/(ah*ah))
print(round(aBMI), ":  wartosc BMI dla osoby", sheet["B3"].value, sheet["C3"].value)
        
bh = sheet["D4"].value
bw = sheet["E4"].value
bBMI = (bw/(bh*bh))
print(round(bBMI), ":  wartosc BMI dla osoby", sheet["B4"].value, sheet["C4"].value)

ch = sheet["D5"].value
cw = sheet["E5"].value
cBMI = (cw/(ch*ch))
print(round(cBMI), ":  wartosc BMI dla osoby", sheet["B5"].value, sheet["C5"].value)

dh = sheet["D6"].value
dw = sheet["E6"].value
dBMI = (dw/(dh*dh))
print(round(dBMI), ":  wartosc BMI dla osoby", sheet["B6"].value, sheet["C6"].value)

eh = sheet["D7"].value
ew = sheet["E7"].value
eBMI = (ew/(eh*eh))
print(round(eBMI), ":  wartosc BMI dla osoby", sheet["B7"].value, sheet["C7"].value)

fh = sheet["D8"].value
fw = sheet["E8"].value
fBMI = (fw/(fh*fh))
print(round(fBMI), ":  wartosc BMI dla osoby", sheet["B8"].value, sheet["C8"].value)

gh = sheet["D9"].value
gw = sheet["E9"].value
gBMI = (gw/(gh*gh))
print(round(gBMI), ":  wartosc BMI dla osoby", sheet["B9"].value, sheet["C9"].value)

hh = sheet["D10"].value
hw = sheet["E10"].value
hBMI = (hw/(hh*hh))
print(round(hBMI), ":  wartosc BMI dla osoby", sheet["B10"].value, sheet["C10"].value)

ih = sheet["D11"].value
iw = sheet["E11"].value
iBMI = (iw/(ih*ih))
print(round(iBMI), ":  wartosc BMI dla osoby", sheet["B11"].value, sheet["C11"].value)

jh = sheet["D12"].value
jw = sheet["E12"].value
jBMI = (jw/(jh*jh))
print(round(jBMI), ":  wartosc BMI dla osoby", sheet["B12"].value, sheet["C12"].value)

kh = sheet["D13"].value
kw = sheet["E13"].value
kBMI = (kw/(kh*kh))
print(round(kBMI), ":  wartosc BMI dla osoby", sheet["B13"].value, sheet["C13"].value)

lh = sheet["D14"].value
lw = sheet["E14"].value
lBMI = (lw/(lh*lh))
print(round(lBMI), ":  wartosc BMI dla osoby", sheet["B14"].value, sheet["C14"].value)

mh = sheet["D15"].value
mw = sheet["E15"].value
mBMI = (mw/(mh*mh))
print(round(mBMI), ":  wartosc BMI dla osoby", sheet["B15"].value, sheet["C15"].value)

nh = sheet["D16"].value
nw = sheet["E16"].value
nBMI = (nw/(nh*nh))
print(round(nBMI), ":  wartosc BMI dla osoby", sheet["B16"].value, sheet["C16"].value)

oh = sheet["D17"].value
ow = sheet["E17"].value
oBMI = (ow/(oh*oh))
print(round(oBMI), ":  wartosc BMI dla osoby", sheet["B17"].value, sheet["C17"].value)

uh = sheet["D18"].value
uw = sheet["E18"].value
uBMI = (uw/(uh*uh))
print(round(uBMI), ":  wartosc BMI dla osoby", sheet["B18"].value, sheet["C18"].value)

ph = sheet["D19"].value
pw = sheet["E19"].value
pBMI = (pw/(ph*ph))
print(round(pBMI), ":  wartosc BMI dla osoby", sheet["B19"].value, sheet["C19"].value)

rh = sheet["D20"].value
rw = sheet["E20"].value
rBMI = (rw/(rh*rh))
print(round(rBMI), ":  wartosc BMI dla osoby", sheet["B20"].value, sheet["C20"].value)

sh = sheet["D21"].value
sw = sheet["E21"].value
sBMI = (sw/(sh*sh))
print(round(sBMI), ":  wartosc BMI dla osoby", sheet["B21"].value, sheet["C21"].value)

th = sheet["D22"].value
tw = sheet["E22"].value
tBMI = (tw/(th*th))
print(round(aBMI), ":  wartosc BMI dla osoby", sheet["B22"].value, sheet["C22"].value)

    